import openpyxl
from app.config import Config

def read_excel():
    """Reads an Excel file with multiple sheets."""
    try:
        wb = openpyxl.load_workbook(Config.EXCEL_FILE)
        data = {}

        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            sheet_data = {}

            for row in sheet.iter_rows(values_only=True):
                if any(row):
                    sheet_data[f"Row {sheet.max_row}"] = row

            data[sheet_name] = sheet_data

        return data
    except Exception as e:
        return {"error": str(e)}

def write_excel():
    """Writes dummy data to a new Excel file."""
    try:
        wb = openpyxl.load_workbook(Config.EXCEL_FILE)
        new_wb = openpyxl.Workbook()

        for sheet_name in wb.sheetnames:
            original_sheet = wb[sheet_name]
            new_sheet = new_wb.create_sheet(title=sheet_name)

            for row in original_sheet.iter_rows():
                for cell in row:
                    new_sheet[cell.coordinate] = cell.value

            # Writing Dummy Data
            new_sheet["A1"] = "Dummy Data 1"
            new_sheet["B2"] = "Dummy Data 2"
            new_sheet["C3"] = "Dummy Data 3"

        if "Sheet" in new_wb.sheetnames:
            std = new_wb["Sheet"]
            new_wb.remove(std)

        new_wb.save(Config.MODIFIED_FILE)
        return {"message": "File written successfully", "file": Config.MODIFIED_FILE}
    except Exception as e:
        return {"error": str(e)}

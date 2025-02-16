import openpyxl
from openpyxl.styles import Border, Side

EXCEL_FILE = "templates/invoice.xlsx"
MODIFIED_FILE = "templates/modified.xlsx"  # New file (downloadable)

def read_excel(file_path=EXCEL_FILE):
    """Reads an Excel file with multiple sheets and returns a dictionary."""
    try:
        wb = openpyxl.load_workbook(file_path)
        data = {}

        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            sheet_data = {}

            # Read non-empty cells
            for row in sheet.iter_rows(values_only=True):
                if any(row):  # Ignore empty rows
                    sheet_data[f"Row {sheet.max_row}"] = row

            data[sheet_name] = sheet_data

        return data
    except Exception as e:
        return {"error": str(e)}


def write_excel(json_data):
    """Writes dummy data to a new Excel file without overwriting the original, preserving borders."""
    try:
        wb = openpyxl.load_workbook(EXCEL_FILE)  # Load existing file
        ws = wb.active  # Select active sheet

        num_new_rows = len(json_data["items"])
        ws.insert_rows(9, amount=num_new_rows)  # Insert rows at row 9

        ws.cell(row=4, column=3, value=json_data["billTo"])
        ws.cell(row=4, column=7, value=json_data["invoice"])
        ws.cell(row=5, column=7, value=json_data["invoiceDate"])
        ws.cell(row=5, column=3, value=json_data["address"])
        ws.cell(row=4, column=4, value="Phone: "+json_data['phone'])
        ws.cell(row=5, column=4, value="Fax: "+json_data['fax'])
        ws.cell(row=6, column=4, value="Email: "+json_data['email'])
        ws.cell(row=7, column=3, value=json_data['invoiceFor'])

        # Define column mapping (B to F)
        columns = ["itemNumber", "description", "qty", "unitPrice", "discount"]
        col_start = 2  # Column B

        # Define border style
        border_style = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # Write new data and apply formula in column G
        for i, row_data in enumerate(json_data["items"], start=9):
            for j, col_name in enumerate(columns):
                cell = ws.cell(row=i, column=col_start + j, value=row_data.get(col_name, ""))
                cell.border = border_style  # Apply border

            # Apply total price formula in column G (7th column)
            qty_cell = f"D{i}"  # Column D (Qty)
            price_cell = f"E{i}"  # Column E (Unit Price)
            discount_cell = f"F{i}"  # Column F (Discount)
            total_cell = ws.cell(row=i, column=7, value=f"=IFERROR({qty_cell}*{price_cell}*(1-{discount_cell}/100), 0)")
            total_cell.border = border_style  # Apply border

        # Total subtotal
        ws.cell(row=num_new_rows+10, column=7, value=f"=SUM(G{num_new_rows}:G{num_new_rows+9})")
        # Tax
        ws.cell(row=num_new_rows + 11, column=7, value=0.18)
        # TaxAmt
        ws.cell(row=num_new_rows + 12, column=7, value=f"=G{num_new_rows+10}*G{num_new_rows+11}")
        # others
        ws.cell(row=num_new_rows + 13, column=7, value=0.0)
        # recieved deposit
        ws.cell(row=num_new_rows + 14, column=7, value=0.0)
        # Payable Amount
        ws.cell(row=num_new_rows + 15, column=7, value=f"=G{num_new_rows+10}+G{num_new_rows+12}+G{num_new_rows+13}+G{num_new_rows+14}")

        wb.save(MODIFIED_FILE)  # Save as a new file to prevent overwriting
        return {"message": "Data written successfully", "file": MODIFIED_FILE}
    except Exception as e:
        return {"error": str(e)}

